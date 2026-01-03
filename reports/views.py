from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from jobs.models import Job
from inventory.models import Radiator


@login_required
def reports_page(request):
    """Display the reports page"""
    return render(request, 'reports/reports_page.html')


@login_required
def download_report(request):
    """Generate and download Excel report with Jobs and Radiators"""
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Get all data
    jobs = Job.objects.all().order_by('-created_at')
    radiators = Radiator.objects.all().order_by('-created_at')
    
    # Define colors
    job_header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")  # Blue
    radiator_header_fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid")  # Green
    job_row_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Light blue
    radiator_row_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Light green
    summary_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Light gray
    
    # Define border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create Summary Sheet
    summary_sheet = wb.create_sheet("Summary", 0)
    summary_sheet.append(["MAGNUM RADIATORS - COMPREHENSIVE REPORT"])
    summary_sheet.append(["Generated on: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_sheet.append([])
    
    # Summary statistics
    summary_data = [
        ["REPORT SUMMARY"],
        [],
        ["JOBS STATISTICS"],
        ["Total Jobs", len(jobs)],
        ["Pending Jobs", len([j for j in jobs if j.status == 'Pending'])],
        ["In Progress Jobs", len([j for j in jobs if j.status == 'In Progress'])],
        ["Completed Jobs", len([j for j in jobs if j.status == 'Completed'])],
        [],
        ["RADIATORS STATISTICS"],
        ["Total Radiators", len(radiators)],
        ["Pending Radiators", len([r for r in radiators if r.status == 'Pending'])],
        ["In Progress Radiators", len([r for r in radiators if r.status == 'In Progress'])],
        ["Completed Radiators", len([r for r in radiators if r.status == 'Completed'])],
        [],
        ["OVERALL STATISTICS"],
        ["Total Records", len(jobs) + len(radiators)],
        ["Total Completed", len([j for j in jobs if j.status == 'Completed']) + len([r for r in radiators if r.status == 'Completed'])],
    ]
    
    for row in summary_data:
        summary_sheet.append(row)
    
    # Style summary sheet
    summary_sheet['A1'].font = Font(bold=True, size=16)
    summary_sheet['A3'].font = Font(bold=True, size=12)
    summary_sheet['A9'].font = Font(bold=True, size=12)
    summary_sheet['A15'].font = Font(bold=True, size=12)
    
    # Auto-adjust column widths for summary
    for col in summary_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        summary_sheet.column_dimensions[column].width = adjusted_width
    
    # Create Combined Sheet
    combined_sheet = wb.create_sheet("Combined", 1)
    
    # Combined headers
    combined_headers = [
        "Type", "Customer Name", "Contact Number", "Vehicle Registration", "Vehicle Make", 
        "Vehicle Model", "Part Type", "Radiator Name", "Work Type", "Status", 
        "Date Received", "Date Completed", "Invoice Number", "Notes", "Created At", "Updated At"
    ]
    combined_sheet.append(combined_headers)
    
    # Style combined headers
    for cell in combined_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Add jobs to combined sheet
    for job in jobs:
        row = [
            "Job",
            job.customer_name,
            job.contact_number,
            job.vehicle_registration,
            job.vehicle_make,
            job.vehicle_model,
            "",  # Part Type
            "",  # Radiator Name
            job.get_work_type_display(),
            job.status,
            job.date_received.strftime("%Y-%m-%d") if job.date_received else "",
            job.date_completed.strftime("%Y-%m-%d") if job.date_completed else "",
            job.invoice_number or "",
            job.notes or "",
            job.created_at.strftime("%Y-%m-%d %H:%M:%S") if job.created_at else "",
            job.updated_at.strftime("%Y-%m-%d %H:%M:%S") if job.updated_at else "",
        ]
        combined_sheet.append(row)
        
        # Style job rows (light blue)
        for cell in combined_sheet[combined_sheet.max_row]:
            cell.fill = job_row_fill
            cell.border = thin_border
    
    # Add radiators to combined sheet
    for radiator in radiators:
        row = [
            "Radiator",
            radiator.customer_name or "",
            radiator.contact_number or "",
            "",  # Vehicle Registration
            "",  # Vehicle Make
            "",  # Vehicle Model
            radiator.get_part_type_display(),
            radiator.name,
            "",  # Work Type
            radiator.status,
            radiator.date_received.strftime("%Y-%m-%d") if radiator.date_received else "",
            radiator.date_completed.strftime("%Y-%m-%d") if radiator.date_completed else "",
            radiator.invoice_number or "",
            radiator.notes or "",
            radiator.created_at.strftime("%Y-%m-%d %H:%M:%S") if radiator.created_at else "",
            radiator.updated_at.strftime("%Y-%m-%d %H:%M:%S") if radiator.updated_at else "",
        ]
        combined_sheet.append(row)
        
        # Style radiator rows (light green)
        for cell in combined_sheet[combined_sheet.max_row]:
            cell.fill = radiator_row_fill
            cell.border = thin_border
    
    # Auto-adjust column widths for combined sheet
    for col in combined_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        combined_sheet.column_dimensions[column].width = adjusted_width
    
    # Create Jobs Sheet
    jobs_sheet = wb.create_sheet("Jobs", 2)
    
    # Jobs headers
    jobs_headers = [
        "Customer Name", "Contact Number", "Vehicle Registration", "Vehicle Make", 
        "Vehicle Model", "Work Type", "Status", "Date Received", "Date Completed", 
        "Invoice Number", "Notes", "Created At", "Updated At"
    ]
    jobs_sheet.append(jobs_headers)
    
    # Style jobs headers
    for cell in jobs_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = job_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Add jobs data
    for job in jobs:
        row = [
            job.customer_name,
            job.contact_number,
            job.vehicle_registration,
            job.vehicle_make,
            job.vehicle_model,
            job.get_work_type_display(),
            job.status,
            job.date_received.strftime("%Y-%m-%d") if job.date_received else "",
            job.date_completed.strftime("%Y-%m-%d") if job.date_completed else "",
            job.invoice_number or "",
            job.notes or "",
            job.created_at.strftime("%Y-%m-%d %H:%M:%S") if job.created_at else "",
            job.updated_at.strftime("%Y-%m-%d %H:%M:%S") if job.updated_at else "",
        ]
        jobs_sheet.append(row)
        
        # Style job rows
        for cell in jobs_sheet[jobs_sheet.max_row]:
            cell.fill = job_row_fill
            cell.border = thin_border
    
    # Auto-adjust column widths for jobs sheet
    for col in jobs_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        jobs_sheet.column_dimensions[column].width = adjusted_width
    
    # Create Radiators Sheet
    radiators_sheet = wb.create_sheet("Radiators", 3)
    
    # Radiators headers
    radiators_headers = [
        "Radiator Name/Model", "Part Type", "Customer Name", "Contact Number", 
        "Status", "Date Received", "Date Completed", "Invoice Number", 
        "Notes", "Created At", "Updated At"
    ]
    radiators_sheet.append(radiators_headers)
    
    # Style radiators headers
    for cell in radiators_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = radiator_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Add radiators data
    for radiator in radiators:
        row = [
            radiator.name,
            radiator.get_part_type_display(),
            radiator.customer_name or "",
            radiator.contact_number or "",
            radiator.status,
            radiator.date_received.strftime("%Y-%m-%d") if radiator.date_received else "",
            radiator.date_completed.strftime("%Y-%m-%d") if radiator.date_completed else "",
            radiator.invoice_number or "",
            radiator.notes or "",
            radiator.created_at.strftime("%Y-%m-%d %H:%M:%S") if radiator.created_at else "",
            radiator.updated_at.strftime("%Y-%m-%d %H:%M:%S") if radiator.updated_at else "",
        ]
        radiators_sheet.append(row)
        
        # Style radiator rows
        for cell in radiators_sheet[radiators_sheet.max_row]:
            cell.fill = radiator_row_fill
            cell.border = thin_border
    
    # Auto-adjust column widths for radiators sheet
    for col in radiators_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        radiators_sheet.column_dimensions[column].width = adjusted_width
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with current date
    filename = f"Magnum_Vehicle/Radiator_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
